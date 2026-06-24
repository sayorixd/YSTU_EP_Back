from src.exceptions import DirectionNotFoundException
from src.directions.model import Direction
from src.direction_map_cors.model import DirectionMapCore
from src.map_cors.model import MapCore
from src.departments.model import Department
from src.disciplines.model import Discipline
from src.discipline_blocks.model import DisciplineBlock
from src.discipline_block_competencies.model import DisciplineBlockCompetency
from src.competencies.model import Competency
from src.competency_groups.model import CompetencyGroup
from src.indicators.model import Indicator
from src.educational_levels.model import EducationalLevel
from src.dependencies import SessionDep
from src.common.excel import \
    set_outer_border_of_cell_range, column_width_from_pixels, row_height_from_pixels, \
    get_height_cell_pixels_after_wrapping_text

from sqlalchemy import select
from fastapi import APIRouter, status, Path, Response
from typing import Annotated
from fastapi.responses import StreamingResponse
from io import StringIO, BytesIO
import csv
from math import floor, ceil
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, DEFAULT_FONT
from openpyxl.styles.borders import Border, Side
import urllib.parse
import base64


router = APIRouter(
    tags=['maps']
)

def map_all_objects(objects_list, key_getter):    
    key_to_object = dict()

    for obj in objects_list:
        key = key_getter(obj)
        if key_to_object.get(key, None) == None:
            key_to_object[key] = []
        key_to_object[key].append(obj)
    
    return key_to_object

def get_unique_items_ordered(objects_list, getter_unique_property=None):
    if getter_unique_property == None:
        getter_unique_property = lambda x: x

    set_objects = set()
    unique_items = []
    for item in objects_list:
        old_length = len(set_objects)
        set_objects.add(getter_unique_property(item))
        if len(set_objects) != old_length:
            unique_items.append(item)
    return unique_items


@router.get(
    '/directions/{direction_id}/competencies_matrix/export/excel',
    responses={
        200: {'description': 'Competencies matrix successfully exported (Excel file)'},
        404: {'description': 'Direction not found'}
    },
    summary='Export the competencies matrix as Excel file'
)
def export_map_excel(direction_id: Annotated[int, Path(gt=0)], session: SessionDep) -> StreamingResponse:
    # V  1. Получить список ядер в направлении подготовки
    # V  2. Получить списки дисциплин, сгрупированные по ядрам
    # V  3. Получить список дисциплин в направлении подготовки
    # V  4. Получить список кафедр по каждой дисциплине
    # V  5. Получить список связей в матрице компетенции
    # V  6. Получить список компетенций в направлении подготовки
    # V  7. Сгруппировать список компетенций по группам компетенций
    # V  8. Отметить плюсиком связи, остальное - пусто
    # X  9. Посчитать итоговое количество плюсиков по дисциплинам
    # X  10. Макет таблицы
    # X  11. Оформление таблицы

    # 1. Получить список дисциплин в направлении подготовки:
    #   1.1. Получить ядра
    #   1.2. Получить дисциплины, добавленные в ядра
    #   1.3. Получить список дисциплин из справочника "Дисциплины"

    direction = session.get(Direction, direction_id)
    if not direction:
        raise DirectionNotFoundException()

    educational_level = session.get(EducationalLevel, direction.educational_level_id)

    # 1. Получить список ядер в направлении подготовки

    direction_map_cores = session.execute(
        select(DirectionMapCore).where(DirectionMapCore.direction_id == direction_id)
    ).scalars().all()

    map_core_ids = [direction_map_core.map_core_id for direction_map_core in direction_map_cores]
    map_cores = session.execute(
        select(MapCore).where(MapCore.id.in_(map_core_ids))
    ).scalars().all()
    map_core_id_to_map_core = {
        map_core.id: map_core
        for map_core in map_cores
    }

    # 2. Получить списки дисциплин, сгруппированные по ядрам
    # 3. Получить список дисциплин в направлении подготовки

    discipline_blocks = session.execute(
        select(DisciplineBlock).where(DisciplineBlock.map_core_id.in_(map_core_ids))
    ).scalars().all()
    
    discipline_blocks_unique_discipline = get_unique_items_ordered(
        discipline_blocks,
        getter_unique_property=lambda x: x.discipline_id
    )

    map_core_ids_of_discipline_blocks = [
        discipline_block.map_core_id
        for discipline_block in discipline_blocks_unique_discipline
    ]

    discipline_ids = [
        discipline_block.discipline_id
        for discipline_block in discipline_blocks_unique_discipline
    ]

    disciplines = session.execute(
        select(Discipline).where(Discipline.id.in_(discipline_ids))
    ).scalars().all()

    disciplines_by_map_core_id = dict()
    for i in range(len(disciplines)):
        map_core_id = map_core_ids_of_discipline_blocks[i]
        if disciplines_by_map_core_id.get(map_core_id, None) == None:
            disciplines_by_map_core_id[map_core_id] = []
        disciplines_by_map_core_id[map_core_id].append(disciplines[i])

    # 4. Получить список кафедр по каждой дисциплине

    department_ids = [discipline.department_id for discipline in disciplines]
    all_departments = session.execute(select(Department)).scalars().all()
    department_id_to_department = {
        department.id: department
        for department in all_departments
    }
    departments = [
        department_id_to_department[department_id]
        for department_id in department_ids
    ]
    
    # 5. Получить список связей в матрице компетенции

    discipline_block_ids = [discipline_block.id for discipline_block in discipline_blocks]
    discipline_block_competencies = session.execute(
        select(DisciplineBlockCompetency)\
        .where(DisciplineBlockCompetency.discipline_block_id.in_(discipline_block_ids))
    ).scalars().all()

    # 6. Получить список компетенций в направлении подготовки

    competency_ids = [
        discipline_block_competency.competency_id
        for discipline_block_competency in get_unique_items_ordered(
            discipline_block_competencies,
            getter_unique_property=lambda x: x.competency_id
        )
    ]
    competencies = session.execute(
        select(Competency).where(Competency.id.in_(competency_ids))
    ).scalars().all()
    competency_id_to_competencies = {
        competency_ids[i]: competencies[i]
        for i in range(len(competency_ids))
    }

    # 7. Сгруппировать компетенции по группам компетенций

    competencies_by_competency_group_id = map_all_objects(
        competencies,
        key_getter=lambda x: x.competency_group_id
    )
    competency_group_ids = [
        key for key in competencies_by_competency_group_id.keys()
    ]
    competency_groups = session.execute(
        select(CompetencyGroup).where(CompetencyGroup.id.in_(competency_group_ids))
    ).scalars().all()
    competency_group_id_to_competency_group = {
        competency_group_ids[i]: competency_groups[i]
        for i in range(len(competency_group_ids))
    }

    # 8. Отметить плюсиком связи, остальное пусто
    
    # 8.1. Преобразовать связи из "Дисциплина ядра" - "Компетенция" в "Дисциплина" - "Компетенция"
    
    discipline_block_id_to_discipline_id = {
        discipline_block.id: discipline_block.discipline_id
        for discipline_block in discipline_blocks
    }

    discipline_competencies = [
        {
            "discipline_id": discipline_block_id_to_discipline_id[discipline_block_competency.discipline_block_id],
            "competency_id": discipline_block_competency.competency_id
        }
        for discipline_block_competency in discipline_block_competencies
    ]

    discipline_ids_by_competency_id = dict()
    for discipline_competency in discipline_competencies:
        key = discipline_competency["competency_id"]
        if discipline_ids_by_competency_id.get(key, None) == None:
            discipline_ids_by_competency_id[key] = set()
        discipline_ids_by_competency_id[key].add(discipline_competency["discipline_id"])

    # 8.2. Проставляем плюсики там, где есть связь

    row_by_competency_id = dict()
    row_by_competency_id = {
        competency.id: [
            "+" if discipline_id in discipline_ids_by_competency_id[competency.id]
            else ""
            for discipline_id in discipline_ids
        ]
        for competency in competencies
    }

    # 9. Посчитать количество плюсиков по дисциплинам
    # 10. Макет таблицы

    wb = Workbook()
    ws = wb.active
    ws.title = direction.name

    # To automatically determine sizes of rows and columns in exported excel file
    # sizes of each character are hardcoded for Arial font in get_text_width function.
    # After changing font family and size of default font you should precompute
    # widths for each character of new font and change hardcoded values
    default_font_id = ("Arial", "11")
    default_font = Font(name="Arial", sz=11)
    for key, value in default_font.__dict__.items():
        setattr(DEFAULT_FONT, key, value)
    title_font = Font(name="Arial", sz=14, b=True)
    header1_font = Font(name="Arial", sz=14)
    map_core_font = Font(name="Arial", sz=14, b=True)
    header2_font = Font(name="Arial", sz=12)

    thin_border = Border(left=Side(style="thin"),
                         top=Side(style="thin"),
                         right=Side(style="thin"),
                         bottom=Side(style="thin"))
    medium_border = Border(left=Side(style="medium"),
                           top=Side(style="medium"),
                           right=Side(style="medium"),
                           bottom=Side(style="medium"))

    color_disciplines = "99cc00"
    color_header_competency_group = "fffd40"
    colors_competency_groups = [
        "fbe5d7",
        "ffe59f",
        "b5cad2"
    ]

    table_width = 3 + len(disciplines)
    table_height = \
        4 + \
        sum([1 + len(c_list) for c_list in competencies_by_competency_group_id.values()]) + \
        1

    rows = [
        [Cell(ws, row=1, column="A", value="") for j in range(table_width)] for i in range(1 + table_height)
    ]

    educational_level_name_in_genetive = educational_level.name_in_genetive.lower()
    direction_name = f"{direction.code} {direction.profile}"
    direction_profile = direction.profile

    i_row = 0
    row = rows[i_row]
    row[0].value = f"Матрица соответствия компетенций и дисциплин по направлению подготовки {direction_name} " + \
                   f"(уровень {educational_level_name_in_genetive})"
    row[0].font = title_font
    row[0].alignment = Alignment(horizontal="left", wrap_text=True)
    row[3].value = f"Направленность (профиль) {direction_profile}"
    row[3].font = title_font
    row[3].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[i_row + 1].height = row_height_from_pixels(71)

    i_row += 1
    row = rows[i_row]
    row[0].value = "№ п/п"
    row[0].font = header1_font
    row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row[1].value = "Компетенции"
    row[1].font = header1_font
    row[1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row[2].value = "Формулировка компетенции"
    row[2].font = header1_font
    row[2].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    i = 3
    for map_core in map_cores:
        count_disciplines_of_map_core = len(disciplines_by_map_core_id[map_core.id])
        row[i].value = map_core.name
        row[i].font = map_core_font
        row[i].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        i += count_disciplines_of_map_core
    ws.row_dimensions[i_row + 1].height = row_height_from_pixels(60)

    i_row += 1
    row = rows[i_row]
    i = 3
    for map_core in map_cores:
        count_disciplines_of_map_core = len(disciplines_by_map_core_id[map_core.id])
        for discipline in disciplines_by_map_core_id[map_core.id]:
            row[i].value = discipline.name
            row[i].alignment = Alignment(horizontal="center", vertical="center", textRotation=90, wrap_text=True)
            row[i].fill = PatternFill(
                start_color=color_disciplines,
                end_color=color_disciplines,
                fill_type="solid"
            )
            i += 1

    disciplines_row_height_pixels = 205
    ws.row_dimensions[i_row + 1].height = row_height_from_pixels(disciplines_row_height_pixels)
    n = 4
    for map_core in map_cores:
        count_disciplines_of_map_core = len(disciplines_by_map_core_id[map_core.id])
        for discipline in disciplines_by_map_core_id[map_core.id]:
            column_width = get_height_cell_pixels_after_wrapping_text(
                discipline.name,
                default_font_id,
                disciplines_row_height_pixels
            ) + 4
            ws.column_dimensions[get_column_letter(n)].width = column_width_from_pixels(column_width)
            n += 1
    
    i_row += 1
    row = rows[i_row]
    for i in range(table_width):
        row[i].fill = PatternFill(
            start_color=color_header_competency_group,
            end_color=color_header_competency_group,
            fill_type="solid"
        )
    row[2].value = "Обеспечивающая кафедра \u2192"
    row[2].alignment = Alignment(horizontal="right", vertical="center")
    i = 3
    for department in departments:
        row[i].value = department.short_name
        row[i].alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        i += 1
    ws.row_dimensions[i_row + 1].height = row_height_from_pixels(60)

    i_row += 1
    row = rows[i_row]
    for i in range(table_width):
        row[i].fill = PatternFill(
            start_color=color_header_competency_group,
            end_color=color_header_competency_group,
            fill_type="solid"
        )
    for i in range(table_width):
        row[i].alignment = Alignment(horizontal="center", vertical="center")
    row[0].value = "1"
    for i in range(2, table_width):
        row[i].value = str(i)
    
    column_code_size_pixels = 51
    column_name_size_pixels = 181
    column_description_size_pixels = 285
    ws.column_dimensions[get_column_letter(1)].width = column_width_from_pixels(column_code_size_pixels)
    ws.column_dimensions[get_column_letter(2)].width = column_width_from_pixels(column_name_size_pixels)
    ws.column_dimensions[get_column_letter(3)].width = column_width_from_pixels(column_description_size_pixels)

    count_plus = [0] * (table_width - 3)
    competency_group_color_i = 0
    for competency_group_id, competencies_of_competency_group in competencies_by_competency_group_id.items():
        competency_group = competency_group_id_to_competency_group[competency_group_id]

        i_row += 1
        row = rows[i_row]
        row[0].value = f"{competency_group.name} ({competency_group.short_name})"
        row[0].alignment = Alignment(horizontal="center", vertical="center")
        for i in range(table_width):
            row[i].fill = PatternFill(
                start_color=color_header_competency_group,
                end_color=color_header_competency_group,
                fill_type="solid"
            )
        
        for competency in competencies_of_competency_group:
            i_row += 1
            row = rows[i_row]
            pattern_fill = PatternFill(
                start_color=colors_competency_groups[competency_group_color_i],
                end_color=colors_competency_groups[competency_group_color_i],
                fill_type="solid"
            )
            row[0].value = competency.code
            row[0].alignment = Alignment(horizontal="center", vertical="center")
            row[0].fill = pattern_fill
            row[1].value = competency.name
            row[1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row[1].fill = pattern_fill
            row[2].value = competency.description
            row[2].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row[2].fill = pattern_fill
            row_plus_or_none = row_by_competency_id[competency.id]
            for i in range(3, table_width):
                row[i].value = row_plus_or_none[i - 3]
                row[i].alignment = Alignment(horizontal="center", vertical="center")
                if row_plus_or_none[i - 3] == "+":
                    count_plus[i - 3] += 1

            name_height = get_height_cell_pixels_after_wrapping_text(
                competency.name,
                default_font_id,
                column_name_size_pixels
            )
            description_height = get_height_cell_pixels_after_wrapping_text(
                competency.description,
                default_font_id,
                column_description_size_pixels
            )
            row_height = max(name_height, description_height)
            ws.row_dimensions[i_row + 1].height = row_height_from_pixels(row_height)

        competency_group_color_i = (competency_group_color_i + 1) % len(colors_competency_groups)

    i_row += 1
    row = rows[i_row]
    row[0].value = "Итого:"
    row[0].alignment = Alignment(horizontal="right")
    for i in range(3, table_width):
        row[i].value = count_plus[i - 3]
        row[i].alignment = Alignment(horizontal="center")

    # Borders

    for i in range(1, len(rows)):
        for j in range(len(rows[0])):
            rows[i][j].border = thin_border

    for i in range(1, 5):
        for j in range(len(rows[0])):
            rows[i][j].border = medium_border

    set_outer_border_of_cell_range(
        rows,
        CellRange(
            min_row=2, min_col=1,
            max_row=1 + table_height, max_col=table_width
        ),
        medium_border
    )

    i_row = table_height
    row = rows[i_row]
    row[0].border = medium_border
    set_outer_border_of_cell_range(
        rows,
        CellRange(
            min_row=i_row + 1, min_col=1,
            max_row=i_row + 1, max_col=table_width
        ),
        medium_border
    )
    
    n = 6
    for competency_group_id, competencies_of_competency_group in competencies_by_competency_group_id.items():
        count_competencies = len(competencies_of_competency_group)
        rows[n - 1][0].border = medium_border
        set_outer_border_of_cell_range(
            rows,
            CellRange(
                min_row=n, min_col=1,
                max_row=n, max_col=table_width
            ),
            medium_border
        )
        n += 1
        for m in range(1, 4):
            set_outer_border_of_cell_range(
                rows,
                CellRange(
                    min_row=n, min_col=m,
                    max_row=n + count_competencies - 1, max_col=m
                ),
                medium_border
            )
        n += count_competencies
        
    # Add rows to worksheet

    for i in range(len(rows)):
        ws.append(rows[i])

    # Merge cells

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=table_width)

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)

    i = 4
    for map_core in map_cores:
        count_disciplines_of_map_core = len(disciplines_by_map_core_id[map_core.id])
        ws.merge_cells(
            start_row=2, start_column=i,
            end_row=2, end_column=i + count_disciplines_of_map_core - 1
        )
        i += count_disciplines_of_map_core
    
    i = 6
    for competency_group_id, competencies_of_competency_group in competencies_by_competency_group_id.items():
        ws.merge_cells(
            start_row=i, start_column=1,
            end_row=i, end_column=3
        )
        i += len(competencies_of_competency_group) + 1

    ws.merge_cells(
        start_row=1 + table_height, start_column=1,
        end_row=1 + table_height, end_column=3
    )

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Матрица компетенций {direction.name}.xlsx"
    encoded_filename = urllib.parse.quote(filename.encode("utf-8"))

    headers = {
        "Content-Disposition": f'attachment; filename=\"competencies_matrix.xlsx\"; filename*=UTF-8\'\'{encoded_filename};',
        "Access-Control-Expose-Headers": "Content-Disposition",
    }

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )
