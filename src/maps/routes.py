from openpyxl.styles import PatternFill
from fastapi import APIRouter, status, Path, Response
from typing import Annotated
from fastapi.responses import StreamingResponse
from io import StringIO, BytesIO
import csv
from openpyxl import Workbook
from sqlalchemy import select
from src.dependencies import MapsServiceDep, SessionDep
from src.control_types.model import ControlType
from .schemas import MapLoad, MapUnload, MapCoreUnload


router = APIRouter(
    tags=['maps']
)


@router.post(
    '/directions/{direction_id}/maps/load',
    status_code=status.HTTP_204_NO_CONTENT,
    responses={
        204: {'description': 'Educational map successfully loaded'},
        404: {'description': 'Direction not found'}
    },
    summary='Load the educational map into the database'
)
def load_map(direction_id: Annotated[int, Path(gt=0)], data: MapLoad, maps_service: MapsServiceDep) -> Response:
    maps_service.load_map(direction_id, data)
    return {'success': 'ok'}


@router.get(
    '/directions/{direction_id}/maps/unload',
    responses={
        200: {'description': 'Educational map successfully unloaded'},
        404: {'description': 'Direction not found'}
    },
    summary='Unload the educational map from the database'
)
def unload_map(direction_id: Annotated[int, Path(gt=0)], maps_service: MapsServiceDep) -> MapUnload:
    return maps_service.unload_map(direction_id)

@router.get(
    '/directions/{direction_id}/maps/export/excel',
    responses={
        200: {'description': 'Educational map successfully exported (Excel file)'},
        404: {'description': 'Direction not found'}
    },
    summary='Export the educational map as Excel file'
)
def export_map_excel(direction_id: Annotated[int, Path(gt=0)], maps_service: MapsServiceDep, session: SessionDep) -> StreamingResponse:
    map_data: MapUnload = maps_service.unload_map(direction_id)
    
    # Динамически получаем все виды контроля
    control_types = session.execute(select(ControlType)).scalars().all()
    primary_types = [ct for ct in control_types if ct.is_primary]
    secondary_types = [ct for ct in control_types if not ct.is_primary]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Educational Plan"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    headers = ['Семестр', 'Ядро', 'Дисциплина', 'Кафедра', 'Зед', 'Зед(час)']
    for pt in primary_types: headers.append(pt.name)
    for st in secondary_types: headers.append(st.name)
    headers += ['Лекционные часы', 'Практические часы', 'Лабораторные часы', 'Сумма часов', 'Разница часов']
    
    ws.append(headers)
    
    for map_core in map_data.map_cors:
        for block in map_core.discipline_blocks:
            zed_hours = block.credit_units * 36
            base_hours = (block.lecture_hours or 0) + (block.practice_hours or 0) + (block.lab_hours or 0)
            
            # ИСПРАВЛЕНО: Сумма часов теперь равна ТОЛЬКО аудиторной нагрузке
            total_hours = base_hours 
            
            # Разница считается между ЗЕ (в часах) и фактической аудиторной нагрузкой
            hours_diff = zed_hours - total_hours 
            
            row = [block.semester_number, map_core.name, block.discipline.name, block.discipline.department.name, block.credit_units, zed_hours]
            
            # Проставляем '+' для основных и дополнительных
            for pt in primary_types:
                # ИСПРАВЛЕНО: сравниваем с block.control_type.id
                row.append('+' if block.control_type and block.control_type.id == pt.id else '')
            for st in secondary_types:
                row.append('+' if st.id in block.secondary_control_type_ids else '')
                
            row += [block.lecture_hours or 0, block.practice_hours or 0, block.lab_hours or 0, total_hours, hours_diff]
            ws.append(row)

            if hours_diff < 0:
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_num).fill = red_fill


    # 5. Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width


    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)


    headers = {
        "Content-Disposition": 'attachment; filename="plan.xlsx"',
        "Access-Control-Expose-Headers": "Content-Disposition",
    }


    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@router.get(
    '/map-cors/{map_core_id}/unload',
    responses={
        200: {'description': 'Map core successfully unloaded'},
        404: {'description': 'Map core not found'}
    },
    summary='Unload the map core from the database'
)
def unload_map_core(map_core_id: Annotated[int, Path(gt=0)], maps_service: MapsServiceDep) -> MapCoreUnload:
    return maps_service.unload_map_core(map_core_id)
