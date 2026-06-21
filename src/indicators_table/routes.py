from src.exceptions import DirectionNotFoundException
from src.directions.model import Direction
from src.direction_map_cors.model import DirectionMapCore
from src.map_cors.model import MapCore
from src.discipline_blocks.model import DisciplineBlock
from src.discipline_block_competencies.model import DisciplineBlockCompetency
from src.competencies.model import Competency
from src.competency_groups.model import CompetencyGroup
from src.indicators.model import Indicator
from src.educational_levels.model import EducationalLevel
from src.dependencies import SessionDep
from src.common.excel import \
    set_outer_border_of_cell_range, column_width_from_pixels, row_height_from_pixels, \
    get_height_cell_pixels_after_wrapping_text

from sqlalchemy import select
from fastapi import APIRouter, status, Path, Response
from typing import Annotated
from fastapi.responses import StreamingResponse
from io import StringIO, BytesIO
import csv
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string
from openpyxl.styles import PatternFill, Font, Alignment, DEFAULT_FONT
from openpyxl.styles.borders import Border, Side
import urllib.parse


router = APIRouter(
    tags=['maps']
)

@router.get(
    '/directions/{direction_id}/indicators_table/export/excel',
    responses={
        200: {'description': 'Indicators table successfully exported (Excel file)'},
        404: {'description': 'Direction not found'}
    },
    summary='Export the indicators table as Excel file'
)
def export_map_excel(direction_id: Annotated[int, Path(gt=0)], session: SessionDep) -> StreamingResponse:
    direction = session.get(Direction, direction_id)
    if not direction:
        raise DirectionNotFoundException()

    all_direction_map_cores = session.execute(select(DirectionMapCore)).scalars().all()
    direction_map_cores = [
        direction_map_core
        for direction_map_core in all_direction_map_cores
        if direction_map_core.direction_id == direction_id
    ]

    map_core_ids = [
        direction_map_core.map_core_id for direction_map_core in direction_map_cores
    ]
    set_map_core_ids = set(map_core_ids)

    all_discipline_blocks = session.execute(select(DisciplineBlock)).scalars().all()
    discipline_block_ids = [
        discipline_block.id
        for discipline_block in all_discipline_blocks
        if discipline_block.map_core_id in set_map_core_ids
    ]
    set_discipline_block_ids = set(discipline_block_ids)

    all_discipline_block_competencies = \
        session.execute(select(DisciplineBlockCompetency)).scalars().all()
    competency_ids = [
        discipline_block_competency.competency_id
        for discipline_block_competency in all_discipline_block_competencies
        if discipline_block_competency.discipline_block_id in set_discipline_block_ids
    ]
    set_competency_ids = set(competency_ids)

    all_competencies = session.execute(select(Competency)).scalars().all()
    competencies = [
        competency
        for competency in all_competencies
        if competency.id in set_competency_ids
    ]

    all_indicators = session.execute(select(Indicator)).scalars().all()
    indicators = [
        indicator
        for indicator in all_indicators
        if indicator.competency_id in set_competency_ids
    ]
    map_competency_to_indicators = dict()
    for competency in competencies:
        map_competency_to_indicators[competency.id] = []
    for indicator in indicators:
        map_competency_to_indicators[indicator.competency_id].append(indicator)

    map_competency_to_indicators_string = dict()
    for competency in competencies:
        map_competency_to_indicators_string[competency.id] = ""
    for indicator in indicators:
        map_competency_to_indicators_string[indicator.competency_id] += \
            indicator.code + " " + indicator.name + "\n"
    for competency_id in map_competency_to_indicators_string.keys():
        if len(map_competency_to_indicators_string[competency_id]) != 0:
            map_competency_to_indicators_string[competency_id] = map_competency_to_indicators_string[competency_id][:-1]

    rows_by_competency_group_id = dict()
    for competency in competencies:
        if rows_by_competency_group_id.get(competency.competency_group_id, None) == None:
            rows_by_competency_group_id[competency.competency_group_id] = []

    for competency in competencies:
        rows_by_competency_group_id[competency.competency_group_id].append([
            competency.code,
            competency.name,
            competency.description,
            map_competency_to_indicators_string[competency.id]
        ])
        #print(rows_by_competency_group_id[competency.competency_group_id])

    competency_group_by_id = {
        competency_group_id: session.get(CompetencyGroup, competency_group_id)
        for competency_group_id in rows_by_competency_group_id.keys()
    }
    
    direction_educational_level = session.get(EducationalLevel, direction.educational_level_id)
    
    direction_educational_level_name = direction_educational_level.name
    direction_name = direction.name
    direction_profile = direction.name[direction.name.index(" ") + 1:]  # TODO: код направления, название, профиль

    

    wb = Workbook()
    ws = wb.active
    ws.title = "Indicators Table"

    default_font_id = ("Arial", "12")
    default_font = Font(name="Arial", sz=12)
    for key, value in default_font.__dict__.items():
        setattr(DEFAULT_FONT, key, value)
    header_font = Font(name="Arial", sz=14, b=True)
    sub_header_font = Font(name="Arial", sz=14)
    table_header_font_id = ("Arial", "12", "bold")
    table_header_font = Font(name="Arial", sz=12, b=True)

    thin_border = Border(left=Side(style="thin"),
                         top=Side(style="thin"),
                         right=Side(style="thin"),
                         bottom=Side(style="thin"))

    color_header_competency_group = "fffd40"
    colors_competency_groups = [
        "fbe5d7",
        "ffe59f",
        "b5cad2"
    ]

    row_n = 1
    row = [f"Компетенции и индикаторы, установленные программой {direction_educational_level_name}"]
    row[0] = Cell(ws, row=1, column="A", value=row[0])
    row[0].font = header_font
    row[0].alignment = Alignment(horizontal="center")
    row[0].border = thin_border
    ws.append(row)
    row_n += 1
    row = [f"Направление: {direction_name}"]
    row[0] = Cell(ws, row=1, column="A", value=row[0])
    row[0].font = sub_header_font
    row[0].alignment = Alignment(horizontal="center")
    row[0].border = thin_border
    ws.append(row)
    row_n += 1
    row = [f"Профиль: {direction_profile}"]
    row[0] = Cell(ws, row=1, column="A", value=row[0])
    row[0].alignment = Alignment(horizontal="center")
    row[0].border = thin_border
    ws.append(row)
    row_n += 1
    row = ["Код", "Группа", "Компетенции", "Индикаторы"]
    for i in range(len(row)):
        row[i] = Cell(ws, row=1, column="A", value=row[i])
        row[i].font = table_header_font
        row[i].alignment = Alignment(horizontal="center")
        row[i].border = thin_border
    ws.append(row)
    row_n += 1
    competency_group_color_i = 0
    for competency_group_id, rows in rows_by_competency_group_id.items():
        competency_group = competency_group_by_id[competency_group_id]
        row = [f"[НАЗВАНИЕ ГРУППЫ КОМПЕТЕНЦИЙ] ({competency_group.name})"]  # TODO: код и название группы компетенций
        row[0] = Cell(ws, row=1, column="A", value=row[0])
        row[0].font = table_header_font
        row[0].alignment = Alignment(horizontal="center")
        row[0].border = thin_border
        row[0].fill = PatternFill(
            start_color=color_header_competency_group,
            end_color=color_header_competency_group,
            fill_type="solid"
        )
        ws.append(row)
        row_n += 1
        for i in range(len(rows)):
            row = rows[i][:]
            for j in range(4):
                row[j] = Cell(ws, row=1, column="A", value=row[j])
                row[j].border = thin_border
            row[0].font = table_header_font
            row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row[0].fill = PatternFill(
                start_color=colors_competency_groups[competency_group_color_i],
                end_color=colors_competency_groups[competency_group_color_i],
                fill_type="solid"
            )
            row[1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row[2].alignment = Alignment(vertical="center", wrap_text=True)
            row[3].alignment = Alignment(vertical="center", wrap_text=True)
            row[3].fill = PatternFill(
                start_color=colors_competency_groups[competency_group_color_i],
                end_color=colors_competency_groups[competency_group_color_i],
                fill_type="solid"
            )
            
            ws.append(row)
            row_n += 1
        competency_group_color_i = (competency_group_color_i + 1) % len(colors_competency_groups)

    competency_code_width_pixels = 80
    competency_name_width_pixels = 207
    competency_description_width_pixels = 400
    indicators_width_pixels = 2000
    ws.column_dimensions[get_column_letter(1)].width = column_width_from_pixels(competency_code_width_pixels)
    ws.column_dimensions[get_column_letter(2)].width = column_width_from_pixels(competency_name_width_pixels)
    ws.column_dimensions[get_column_letter(3)].width = column_width_from_pixels(competency_description_width_pixels)
    ws.column_dimensions[get_column_letter(4)].width = column_width_from_pixels(indicators_width_pixels)
    n = 5
    for competency_group_id, rows in rows_by_competency_group_id.items():
        n += 1
        for i in range(len(rows)):
            row = rows[i][:]

            competency_description = str(row[2])
            indicators = str(row[3])

            competency_description_height = get_height_cell_pixels_after_wrapping_text(
                competency_description,
                default_font_id,
                competency_description_width_pixels
            )
            indicators_height = get_height_cell_pixels_after_wrapping_text(
                indicators,
                default_font_id,
                indicators_width_pixels
            )
            max_height = max(competency_description_height, indicators_height)

            ws.row_dimensions[n].height = row_height_from_pixels(max_height)
            n += 1
    
    i = 0
    for competency_group_id, rows in rows_by_competency_group_id.items():
        ws.merge_cells(
            start_row=5 + i, start_column=1,
            end_row=5 + i, end_column=4
        )
        i += len(rows) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Перечень компетенций {direction.name}.xlsx"
    encoded_filename = urllib.parse.quote(filename.encode("utf-8"))

    headers = {
        "Content-Disposition": f'attachment; filename=\"indicators_table.xlsx\"; filename=UTF-8\'\'{encoded_filename}',
        "Access-Control-Expose-Headers": "Content-Disposition",
    }


    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )
